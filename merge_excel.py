import argparse
import glob
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import queue
import threading
from collections import OrderedDict
import time

# ----------------------
# Arguments
# ----------------------
parser = argparse.ArgumentParser()
parser.add_argument(
    "--source-info",
    action="store_true",
    help="Include source information"
)
parser.add_argument(
    "--keep-duplicates",
    action="store_true",
    help="Keep duplicate rows instead of removing them"
)
parser.add_argument(
    "--threads",
    type=int,
    default=8,
    help="Number of reading threads"
)
parser.add_argument(
    "--batch-size",
    type=int,
    default=1000,
    help="Number of rows to process in batch"
)
args = parser.parse_args()

print("Source info enabled:", args.source_info)
print("Keep duplicates:", args.keep_duplicates)
print("Threads:", args.threads)
print("Batch size:", args.batch_size)

# ----------------------
# Collect files
# ----------------------
files = glob.glob("files/*.xlsx")
print(f"Found {len(files)} files to process")

# ----------------------
# Output workbook (streaming)
# ----------------------
wb_out = Workbook(write_only=True)
ws_out = wb_out.create_sheet("Merged")

# ----------------------
# Global header management
# ----------------------
header_lock = threading.Lock()
all_headers = OrderedDict()
header_ready = threading.Event()
seen_rows = set() if not args.keep_duplicates else None

def extract_data_from_chartsheet(ws):
    """Extract data from chartsheet by checking for any available data"""
    rows = []

    # Try to get data from chart titles or any text elements
    try:
        # Chartsheets might have a title or some text content
        if hasattr(ws, 'title'):
            rows.append(['Chart_Title', ws.title])

        # Some chartsheets might have embedded data tables
        # This is a best-effort approach to extract any available data
        if hasattr(ws, 'charts') and ws.charts:
            for i, chart in enumerate(ws.charts):
                # Try to get chart title
                if hasattr(chart, 'title') and chart.title and hasattr(chart.title, 'text'):
                    rows.append([f'Chart_{i}_Title', str(chart.title.text)])

                # Try to get series names if available
                if hasattr(chart, 'series') and chart.series:
                    for j, series in enumerate(chart.series):
                        series_name = getattr(series, 'name', f'Series_{j}')
                        rows.append([f'Chart_{i}_Series_{j}', str(series_name)])
    except Exception as e:
        # If we can't extract data, just return empty
        pass

    return rows

def process_sheet(ws, sheet_name, f):
    """Process any type of sheet (worksheet or chartsheet)"""
    rows = []

    # Regular worksheet
    if hasattr(ws, 'iter_rows'):
        try:
            for row in ws.iter_rows(values_only=True):
                if row and any(cell is not None for cell in row):
                    rows.append([str(cell).strip() if cell is not None else "" for cell in row])
        except Exception as e:
            print(f"Warning: Could not read worksheet '{sheet_name}' in file '{f}': {e}")

    # Chartsheet - try to extract any available data
    elif hasattr(ws, 'charts'):
        try:
            chart_data = extract_data_from_chartsheet(ws)
            if chart_data:
                rows.extend(chart_data)
                # Add a simple header if we found data
                if rows and len(rows[0]) == 2:
                    rows.insert(0, ['Chart_Element', 'Value'])
        except Exception as e:
            print(f"Info: Limited data available in chartsheet '{sheet_name}' in file '{f}'")

    # Hidden sheets
    elif hasattr(ws, 'sheet_state') and ws.sheet_state == 'hidden':
        print(f"Info: Skipping hidden sheet '{sheet_name}' in file '{f}'")
        return []

    return rows

# ----------------------
# Thread-safe queue for rows with batching
# ----------------------
row_queue = queue.Queue(maxsize=5000)

def file_reader(f):
    try:
        wb_in = load_workbook(f, read_only=True, data_only=True)
        source_headers = ["source_file", "source_sheet"] if args.source_info else []

        for sheet_name in wb_in.sheetnames:
            ws_in = wb_in[sheet_name]

            # Process all types of sheets
            rows = process_sheet(ws_in, sheet_name, f)

            if not rows:
                continue

            first_row = rows[0]
            file_headers = [str(h).strip() if h is not None else "" for h in first_row]

            # Cache header mapping for this file
            header_mapping = []
            for i, header in enumerate(file_headers):
                with header_lock:
                    if header not in all_headers:
                        all_headers[header] = len(all_headers)
                    header_mapping.append((i, all_headers[header]))

            # Add source headers to mapping
            source_mapping = []
            for header in source_headers:
                with header_lock:
                    if header not in all_headers:
                        all_headers[header] = len(all_headers)
                    source_mapping.append(all_headers[header])

            # Send header if first file
            if not header_ready.is_set():
                with header_lock:
                    full_header_row = [""] * len(all_headers)
                    for header, idx in all_headers.items():
                        full_header_row[idx] = header
                    row_queue.put(("header", full_header_row))
                    header_ready.set()

            # Process data rows in batches
            batch = []
            for row in rows[1:]:  # Skip header row
                if not row:
                    continue

                # Create ordered row with pre-allocated list
                ordered_row = [""] * len(all_headers)

                # Map file data
                for src_idx, dst_idx in header_mapping:
                    if src_idx < len(row) and row[src_idx] is not None:
                        ordered_row[dst_idx] = str(row[src_idx]).strip()

                # Add source info
                if args.source_info:
                    for i, dst_idx in enumerate(source_mapping):
                        if i == 0:
                            ordered_row[dst_idx] = f
                        elif i == 1:
                            ordered_row[dst_idx] = sheet_name

                # Check duplicates
                row_tuple = tuple(ordered_row)
                if args.keep_duplicates or row_tuple not in seen_rows:
                    if not args.keep_duplicates:
                        seen_rows.add(row_tuple)
                    batch.append(ordered_row)

                    # Send batch when full
                    if len(batch) >= args.batch_size:
                        row_queue.put(("batch", batch))
                        batch = []

            # Send remaining rows in batch
            if batch:
                row_queue.put(("batch", batch))

    except Exception as e:
        print(f"Error processing file {f}: {e}")
    finally:
        if 'wb_in' in locals():
            wb_in.close()
    return f

# ----------------------
# Writer thread with batching
# ----------------------
def writer():
    processed = 0
    start_time = time.time()

    while True:
        item = row_queue.get()
        if item == "DONE":
            break

        typ, data = item
        if typ == "batch":
            for row in data:
                ws_out.append(row)
                processed += 1

                # Progress reporting
                if processed % 10000 == 0:
                    elapsed = time.time() - start_time
                    print(f"Processed {processed} rows in {elapsed:.2f}s ({processed/elapsed:.0f} rows/s)")
        else:
            ws_out.append(data)

        row_queue.task_done()

    print(f"Total rows written: {processed}")

# ----------------------
# Start writer thread
# ----------------------
writer_thread = threading.Thread(target=writer)
writer_thread.start()

# ----------------------
# Start reading threads with progress
# ----------------------
start_time = time.time()
with ThreadPoolExecutor(max_workers=args.threads) as executor:
    futures = {executor.submit(file_reader, f): f for f in files}

    for future in tqdm(as_completed(futures), total=len(futures), desc="Processing files"):
        try:
            future.result()
        except Exception as e:
            print(f"Error: {e}")

# ----------------------
# Finish
# ----------------------
row_queue.put("DONE")
writer_thread.join()

total_time = time.time() - start_time
print(f"Processing completed in {total_time:.2f} seconds")

wb_out.save("result/merged.xlsx")
print("Done → result/merged.xlsx")
